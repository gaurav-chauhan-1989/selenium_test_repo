import openpyxl

wb = openpyxl.load_workbook("E:\Sheet\wrt.xlsx")
sh = wb.active
a=[]
for i in range(1, sh.max_row+1):
    for j in range(1, sh.max_column+1):
        a.append(sh.cell(row=i, column=j).value)
print(a)